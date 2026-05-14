"""Build FYDP / LPC wide Excel workbooks that match ``_detect_lpc_wide_layout`` / ingest column names."""

from __future__ import annotations

import io
import re
from typing import Iterable, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

_MATRIX_REMARKS_SPLIT = '\n\n— Remarks —\n'


def _hierarchy_labels(record):
    pap = record.indicator.pap
    strat = pap.parent if pap else None
    out = strat.parent if strat else None
    outcome = (out.name if out else '').strip() or 'Outcome'
    strategy = (strat.name if strat else '').strip() or 'Strategy'
    pap_name = (pap.name if pap else '').strip() or 'Program / PAP'
    return outcome, strategy, pap_name


def _split_pap_subpap(pap_name: str) -> tuple[str, str]:
    if not pap_name:
        return '', ''
    if ' — ' in pap_name:
        left, right = pap_name.split(' — ', 1)
        return left.strip(), right.strip()
    return pap_name.strip(), ''


def _split_accomplishment_remarks(raw: str) -> tuple[str, str]:
    raw = (raw or '').strip()
    if _MATRIX_REMARKS_SPLIT in raw:
        acmp, _, rem = raw.partition(_MATRIX_REMARKS_SPLIT)
        return acmp.strip(), rem.strip()
    return raw, ''


def _format_num(val) -> str:
    if val is None:
        return ''
    try:
        if float(val).is_integer():
            return str(int(float(val)))
        return str(round(float(val), 4)).rstrip('0').rstrip('.')
    except (TypeError, ValueError):
        return str(val)


def _status_cell(record) -> str:
    exp = (getattr(record, 'explicit_status', None) or '').strip().upper()
    if exp in ('MET', 'UNMET'):
        return exp
    st = (record.status or 'PENDING').strip().upper()
    if st in ('MET', 'UNMET'):
        return st
    return 'PENDING'


def lpc_wide_header_labels(report_year: int) -> List[str]:
    """Column titles aligned with institutional LPC templates and ``_map_lpc_wide_columns``."""
    fy = int(report_year)
    return [
        'Outcome',
        'Strategy (Based on FYDP):',
        'Program/Activity/Project (Based on FYDP):',
        'Sub PAP (Based on FYDP):',
        'Performance Indicators (based on the Cascaded Template):',
        'Performance Indicators (based on the submitted Operational Plan):',
        'Concerned Office / Campus:',
        'Q1 Target (Based on Annual Operational Plan):',
        'Q2 Target (Based on Annual Operational Plan):',
        'Q3 Target (Based on Annual Operational Plan):',
        'Q4 Target (Based on Annual Operational Plan):',
        'Annual Quantifiable Target (Based on Annual Operational Plan):',
        'Accomplishment to Date (Sum of Quarterly Accomplishments):',
        'Variance to Date (Quarterly vs Targets):',
        'Status of Accomplishment (MET or UNMET):',
        f'FY {fy} Overall Details of Accomplishments:',
    ]


def build_lpc_wide_workbook_bytes(
    records: Iterable,
    *,
    dev_area_banner: str,
    report_year: int,
    view_all_quarters: bool,
) -> bytes:
    """One data row per ``PerformanceRecord``; hierarchy in columns 1–4; indicator in col 6 (submitted)."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'LPC Export'

    hdr_fill = PatternFill('solid', fgColor='FCE5CD')
    hdr_font = Font(bold=True, size=10)
    wrap = Alignment(wrap_text=True, vertical='top')

    banner = (dev_area_banner or '').strip()
    if not banner:
        banner = 'DEVELOPMENT AREA: (not specified)'
    if not re.match(r'(?i)^development\s+area\s*:', banner):
        banner = f'DEVELOPMENT AREA: {banner}'
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=16)
    c1 = ws.cell(row=1, column=1, value=banner)
    c1.font = Font(bold=True, size=11, color='C00000')
    c1.alignment = wrap

    headers = lpc_wide_header_labels(report_year)
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=title)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

    row = 3
    fy = int(report_year)
    for rec in records:
        outcome, strategy, pap = _hierarchy_labels(rec)
        pap_main, sub_pap = _split_pap_subpap(pap)
        ind = (rec.indicator.description or '').strip()
        office = ''
        if rec.indicator.pap and rec.indicator.pap.office:
            office = (rec.indicator.pap.office.name or '').strip()

        acmp, remarks = _split_accomplishment_remarks(rec.raw_actual_text or '')
        annual = _format_num(rec.target_value)
        q_targets = ['', '', '', '']
        qn = int(rec.quarter or 0)
        if 1 <= qn <= 4:
            q_targets[qn - 1] = annual or (rec.indicator.target_text or '').strip()
        status = _status_cell(rec)
        var_txt = (rec.variance_text or '').strip()
        fy_details = acmp
        if remarks:
            fy_details = f'{acmp}\n\nBest practices / notes:\n{remarks}' if acmp else remarks

        line = [
            outcome,
            strategy,
            pap_main,
            sub_pap,
            '',
            ind,
            office,
            q_targets[0],
            q_targets[1],
            q_targets[2],
            q_targets[3],
            annual,
            acmp,
            var_txt,
            status,
            fy_details,
        ]
        for col_idx, val in enumerate(line, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.alignment = wrap
        row += 1

    from openpyxl.utils import get_column_letter

    widths = (28, 36, 40, 24, 20, 44, 28, 22, 22, 22, 22, 16, 36, 18, 18, 50)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    bio = io.BytesIO()
    wb.save(bio)
    wb.close()
    return bio.getvalue()
